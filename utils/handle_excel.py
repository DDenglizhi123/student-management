import  openpyxl

class ReadExcel():
    def __init__(self,file_path:str):
        self.wb = openpyxl.load_workbook(file_path)
        self.ws = self.wb.active

    def get_data(self) -> list:
        data:list = []
        for row in self.ws.iter_rows(): # type: ignore
            row_value:list = []
            for cell in row:
                row_value.append(cell.value)
            data.append(row_value)
        return data


class WriteExcel():
    
    def __init__(self,file_path,data) -> None:
        self.file_path = file_path
        self.data= data
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        
    def write_to_excel(self):
        for row in self.data:
            self.ws.append(row) # type: ignore
        self.wb.save(self.file_path)
    
if __name__ == '__main__':
    # read_excel_obj:object = ReadExcel('/Users/dlz/Documents/student-management/1年1班学生信息.xlsx')
    # result = read_excel_obj.get_data()
    # print(result)
    data = []
    write_excel_obj = WriteExcel('/Users/dlz/Documents/student-management/测试一下.xlsx', data)
    write_excel_obj.write_to_excel()